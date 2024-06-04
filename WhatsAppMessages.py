import pywhatkit as pw #libreria para interacuar con whatsapp
import pyautogui  #libreria para controlar el teclado
import time #libreria para obtener la hora
import openpyxl as ex #libreria para excel
import sys

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QMessageBox, QWidget, QPushButton, QMainWindow, QTextEdit, QLabel, QTableWidget, QTableWidgetItem, QApplication
from PyQt5.QtGui import QIcon, QFont, QCursor

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # Establecer el tamaño y título de la ventana
        self.setGeometry(100, 100, 925, 550)
        self.setWindowTitle('WhatsAppMessages')
        self.setWindowIcon(QIcon('python.png'))
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowMaximizeButtonHint) #Impedir que se maximice la ventana
        #self.setWindowFlags(self.windowFlags() & ~Qt.WindowMinimizeButtonHint & ~Qt.WindowMaximizeButtonHint)

        # Crear un botón y establecer su posición y tamaño
        self.btn_cargar = QPushButton('Cargar Archivo', self)
        self.btn_cargar.setGeometry(800, 80, 100, 50)
        self.btn_cargar.setIcon(QIcon("carga.png"))
        # Conectar la señal clicked del botón al método on_button_click
        self.btn_cargar.setToolTip("Cargar el archivo de Excel") #Anade una etiqueta para cuando el cursor este sobre el boton
        self.btn_cargar.setCursor(QCursor(Qt.PointingHandCursor)) #Modifica la forma del cursor sobre el boton
        self.btn_cargar.clicked.connect(self.btn_cargarArchivo)

        self.button = QPushButton("Enviar", self)
        self.button.setIcon(QIcon('whatsapp.png')) #agrega un icono al boton
        self.button.setGeometry(800, 150, 100, 50) # X, Y, W, H
        self.button.setToolTip("Iniciar el envio de mensajes")
        self.button.setCursor(QCursor(Qt.PointingHandCursor))
        self.button.clicked.connect(self.btn_enviar)

        global text_edit
        mensaje = f"Este es el mensaje que recibiran los clientes"
        self.text_edit = QTextEdit(mensaje, self)
        self.text_edit.setGeometry(50, 80, 700, 120)
        font= QFont("Arial", 14)
        self.text_edit.setFont(font)

        self.label1 = QLabel("Escribe aqui el mensaje que enviaras", self)
        self.label1.setGeometry(50, 50, 200, 20)


        self.tabla = QTableWidget(200, 5,self)
        #El espacio en blanco es para que se distancie en la tabla
        self.cabecera = ('Cedula                ','Nombre                                             ',
                         'Telefono           ', 'Monto Atraso', 'Dias de Atraso')
        self.tabla.setHorizontalHeaderLabels(self.cabecera)
        self.tabla.resizeColumnsToContents()
        self.tabla.setGeometry(50, 225, 700, 300)

    def btn_enviar(self):
        try:
            msg = QMessageBox()
            msg.information(self, "Informacion", "Iniciando el envio de mensajes. Por favor espere...")
            self.sentmesagge(numero, mensaje)

        except(Exception):
            arch_on = QMessageBox()
            arch_on.information(self, "Advertencia", "Ha ocurrido un error.")

    def sentmesagge(self, numero,mensaje):
        # Obteniendo la hora exacta y un minuto mas
        hora = time.strftime("%H:%M:%S")  # Formato de 24 horas
        hour = int(hora[0:2])
        minute = int(hora[3:5]) + 1

        pw.sendwhatmsg(numero, mensaje, hour, minute)
        pyautogui.press('enter') #enviar el mensaje
        time.sleep(5)  # Espera 5 segundos para que el mensaje se envie correctamente
        pyautogui.hotkey('ctrl', 'w')
        pyautogui.press('enter')

    def btn_cargarArchivo(self):
        global lista, contador, numero, nombre, mensaje, atraso, archivos_on
        archivos_on = True
        libro = ex.load_workbook("DDBB.xlsx")
        lista = libro.active

        for i in range(lista.max_row + 1):
            if i <= 1: continue  # esta linea sirve para que no se impriman los titulos en el archivo excel

            #Elegir el numero de telefono de la columna 5 o la 6 en caso de que no haya un numero en la no. 5
            if str(lista.cell(row=i, column=3).value) == ".":
                numero = f"+1{str(lista.cell(row=i, column=4).value)}".replace("-", "") #Limpiar el numero de guiones
            else:
                numero = f"+1{str(lista.cell(row=i, column=3).value)}".replace("-", "")

            atraso = lista.cell(row=i, column=5).value
            nombre = lista.cell(row=i, column=2).value
            mensaje = self.text_edit.toPlainText()
            monto = lista.cell(row=i, column=6).value
            cedula = lista.cell(row=i, column=1).value

            x = i -2 #Esta linea sirve para poder añadir los datos a la tabla sin dejar espacios en blanco
            #insertando la cedula en la tabla
            dato = QTableWidgetItem(f"{cedula}")
            self.tabla.setItem(x, 0, dato)

            #insertando el nombre a la tabla
            dato = QTableWidgetItem(nombre)
            self.tabla.setItem(x, 1, dato)

            # insertando el numero a la tabla
            dato = QTableWidgetItem(numero)
            self.tabla.setItem(x, 2, dato)

            # insertando el monto en atraso a la tabla
            dato = QTableWidgetItem(f"{monto}")
            self.tabla.setItem(x, 3, dato)

            # insertando dias de atraso a la tabla
            dato = QTableWidgetItem(f"{atraso}")
            self.tabla.setItem(x, 4, dato)

        #Mensaje de que los datos han sido cargados
        msg = QMessageBox()
        msg.information(self, "Informacion", "Datos Cargados con Exito")


if __name__ == "__main__":
    #crear la aplicacion
    app = QApplication(sys.argv)

    #crear la ventana
    mainWindow = MainWindow()
    mainWindow.show()

    # Iniciar el bucle de eventos de la aplicación
    sys.exit(app.exec_())


"""Notas y Errores
Nota 1: tener pendiente el .value para obtener los valores correctamente

Error1: Al precionar el boton de 'cargar archivo' el programa se cerraba, esto debido a que no se asignaba ningun valor
a la variable 'nombre', ni a la variable 'atraso' en 'mensaje'

RECORDATORIOS:
Dias en mora casilla 9 en la DDBB
"""
